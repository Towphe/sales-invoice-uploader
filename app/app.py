from PyQt6.QtCore import QSize, Qt
from PyQt6.QtGui import QMouseEvent
from PyQt6.QtWidgets import QApplication, QMainWindow, QComboBox, QPushButton, QVBoxLayout, QWidget, QLabel, QLineEdit, QFileDialog, QDialog
from openpyxl import Workbook, worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import functools
import sys
from util import create_template, isValidInt
import time

class UploadButton(QPushButton):
    def get_filename(self):
        print(self.f_name)
        return self.f_name
    
    # def return_func_wrapper(self):
    #     return functools.partial(self.on_clicked)

    def upload(self, main, f_name):
        self.f_name = QFileDialog.getOpenFileName(self,
                                                  "Open File",
                                                  "${self.previous_dir}",
                                                  "Excel Workbooks (*.xlsx)") # filter for excel worksheets later
        # this stores file
        # try to get root name

        f_name_split = self.f_name[0].split("/")
        f_name.setText("File name: " + f_name_split[-1])
        main.file_dict[self.id[7:len(self.id)]] = self.f_name
        
    def __init__(self, label):
        super().__init__(label)
        self.id = label
        self.previous_dir = "${HOME}"
        self.f_name = ""

class MainWindow(QMainWindow):
    # store directories
    def store_si(self):
        # validate later
        self.si_no = self.si_input.text()
    
    def store_year(self):
        self.year = self.year_input.text()

    def store_file_name(self):
        self.file_name = self.file_name_input.text()
    
    def dlg_close(dlg:QPushButton):
        dlg.close()

    def submit(self):
        self.loading_label.setText("Loading...")
        QApplication.processEvents()

        status = "Convert successful"

        # validate text inputs
        self.file_name = self.file_name_input.text()
        self.si_no = self.si_input.text()
        self.si_month = self.month_selection.currentText()
        self.year = self.year_input.text()
        if (not isValidInt(self.year)):
            # mark as unsuccessful
            is_valid = False
            message = "Year"

        # then call util function here
        output = create_template(self.si_no, self.si_month, int(self.year), self.file_dict.get('Lazada SOA File')[0], self.file_dict.get('QNE Sales Order Report File')[0], self.file_dict.get('QNE Sales Invoice Report File')[0], self.file_dict.get('QNE Sales Order Register File')[0])
        
        dlg = QDialog(self)
        dlg.setWindowTitle("Success")

        if type(output) is tuple:
            message = QLabel(output[0])
        else:
            # get pandas file
            # edit a bit
            # open directory prompt
            wb = Workbook()
            ws = wb.active

            for r in dataframe_to_rows(output, index=False, header=True):
                ws.append(r)

            # color first row accordingly
            orange_fill = PatternFill(start_color='D2691E', fill_type="solid")
            for col in ws['A':'AH']:
                col[0].fill = orange_fill

            grey_fill = PatternFill(start_color='808080', fill_type="solid")
            for i in range(1,output.shape[0]):
                ws['AI' + str(i)].fill = grey_fill
            
            #ws['AI1'].fill = PatternFill(start_color='808080', fill_type="solid")
            
            blue_fill = PatternFill(start_color='00BFFF', fill_type="solid")
            for col in ws['AJ':'BU']:
                col[0].fill = blue_fill
            
            ws['AI1'] = ""

            for i in range(2, output.shape[0] + 2):
                ws.cell(row=i, column=2).number_format = 'mm-dd-yyyy'
                ws.cell(row=i, column=3).number_format = 'mm-dd-yyyy'
                ws.cell(row=i, column=5).number_format = 'mm-dd-yyyy'
                ws.cell(row=i, column=9).number_format = 'mm-dd-yyyy'
                ws.cell(row=i,column=15).number_format = '0'
                ws.cell(row=i,column=48).number_format = '0'
                ws.cell(row=i,column=50).number_format = '0.00'

            # save directory
            # save file to directory
            self.file_dir = QFileDialog.getExistingDirectory(self, "Save File")
            wb.save(self.file_dir + "/" + self.file_name + ".xlsx")
            
            message = QLabel(status)

        self.loading_label.setText("")

        dlg.close_button = QPushButton("Close")
        dlg.close_button.clicked.connect(dlg.close)

        dlg_layout = QVBoxLayout()

        dlg_layout.addWidget(message)
        dlg_layout.addWidget(dlg.close_button)
        dlg.setLayout(dlg_layout)

        dlg.exec()
    
    def __init__(self):
        super().__init__()

        #dialog_wrap = functools.partial(self.on_clicked, self)
        self.f_name = ""
        self.file_dict = dict()

        # set window defaults
        self.setWindowTitle("Sales Invoice Uploader")
        self.setFixedSize(QSize(450, 570))

        self.file_name_label = QLabel("Filename")
        self.file_name_input = QLineEdit()
        self.file_name_input.textChanged.connect(self.store_file_name)
        
        self.si_label = QLabel("SI Number Start")
        self.si_input = QLineEdit()
        self.si_input.textChanged.connect(self.store_si)

        # add elements

        self.month_label = QLabel("Month")
        self.month_selection = QComboBox()
        self.month_selection.addItems(['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'])

        self.year_label = QLabel("Year")
        self.year_input = QLineEdit()
        self.year_input.textChanged.connect(self.store_year)

        self.sorep_label = QLabel("QNE Sales Order Report File")
        self.sorep_upload = UploadButton("Upload QNE Sales Order Report File")
        self.sorep_fname = QLabel("")
        self.sorep_upload.clicked.connect(functools.partial(self.sorep_upload.upload, self, self.sorep_fname))
        # receive file

        self.sirep_label = QLabel("QNE Sales Invoice Report File")
        self.sirep_upload = UploadButton("Upload QNE Sales Invoice Report File")
        self.sirep_fname = QLabel("")
        self.sirep_upload.clicked.connect(functools.partial(self.sirep_upload.upload, self, self.sirep_fname))

        self.soreg_label = QLabel("QNE Sales Order Register File")
        self.soreg_upload = UploadButton("Upload QNE Sales Order Register File")
        self.soreg_fname = QLabel("")
        self.soreg_upload.clicked.connect(functools.partial(self.soreg_upload.upload, self, self.soreg_fname))

        self.soa_label = QLabel("Lazada SOA File")
        self.soa_upload = UploadButton("Upload Lazada SOA File")
        self.soa_upload_fname = QLabel("")
        self.soa_upload.clicked.connect(functools.partial(self.soa_upload.upload, self, self.soa_upload_fname))
        
        self.loading_label = QLabel("")

        submit_button = QPushButton("Generate")
        submit_button.clicked.connect(self.submit)
        
        

        layout = QVBoxLayout()

        layout.addWidget(self.file_name_label)
        layout.addWidget(self.file_name_input)

        layout.addWidget(self.si_label)
        layout.addWidget(self.si_input)

        layout.addWidget(self.month_label)
        layout.addWidget(self.month_selection)

        layout.addWidget(self.year_label)
        layout.addWidget(self.year_input)

        layout.addWidget(self.soa_label)
        layout.addWidget(self.soa_upload)
        layout.addWidget(self.soa_upload_fname)

        layout.addWidget(self.sorep_label)
        layout.addWidget(self.sorep_upload)
        layout.addWidget(self.sorep_fname)

        layout.addWidget(self.sirep_label)
        layout.addWidget(self.sirep_upload)
        layout.addWidget(self.sirep_fname)

        layout.addWidget(self.soreg_label)
        layout.addWidget(self.soreg_upload)
        layout.addWidget(self.soreg_fname)

        layout.addWidget(submit_button)
        layout.addWidget(self.loading_label)
        
        container = QWidget()
        container.setLayout(layout)
        self.setMenuWidget(container)

# Create instance of Qt application
app = QApplication(sys.argv)

# Create window instance
window = MainWindow()
window.show() # show window

# start application loop
app.exec()